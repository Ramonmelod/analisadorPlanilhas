import pandas as pd
from openpyxl import load_workbook  # load_workbook permite abrir e carregar um arquivo Excel existente para manipulação.
import re
import os  # biblioteca usada para utilização do metodo de listagem do nome dos arquivos

caminhopasta = "C:\\Users\\ramon\\OneDrive\\Documentos\\GitHub\\analisadorPlanilhas\\planilhas\\"  # necessário implementar container para que o caminho da pasta se mantenha o mesmo idependente da máquina que o código roda
#caminhopasta = "C:\\Users\\ramon\\Documents\\GitHub\\analisadorPlanilhas\\planilhas\\" # pcTrabalho
colecaoPlanilhas =[]

nomeArquivos = os.listdir(caminhopasta) # Lista todos os arquivos na pasta e armazena em uma array

# print(nomeArquivos[16])
# planilha = pd.read_excel(caminhopasta + nomeArquivos[16], sheet_name=0)
# print("Nomes das colunas:", planilha.columns)


#print(planilha["Aparelho gela?"])


for i in range(0, 46):
   #print(i)
   #print(nomeArquivos[i])
   planilhaSetor = pd.read_excel(caminhopasta + nomeArquivos[i], sheet_name=0)
   colecaoPlanilhas.append(planilhaSetor)

regex_padrao = r'Aparelho gela\??|aparelho gela\?|Aparelho gelsa|aparelho gela'



planilhaInicial = colecaoPlanilhas[0]

print('-------------início for---------------')    
for i in range(1,10): 
   print(nomeArquivos[i - 1],"i:  " , i - 1)
   #print(colecaoPlanilhas[i]['Aparelho gela?'])
   #print(colecaoPlanilhas[i].columns)
   print("planilha: ", nomeArquivos[i], "  +  ", nomeArquivos[i-1])
   planilhaInicial = pd.concat( [planilhaInicial,colecaoPlanilhas[i]] )  

   # if colecaoPlanilhas[i].columns.equals(colecaoPlanilhas[i-1].columns):
   #    print("planilha: ", nomeArquivos[i], "  +  ", nomeArquivos[i-1])
   #    planilhaInicial = pd.concat( [planilhaInicial,colecaoPlanilhas[i]] )  
      

      

   print('-------------fim planilha---------------')

planilhaInicial.to_excel("C:\\Users\\ramon\\OneDrive\\Documentos\GitHub\\analisadorPlanilhas\\planilha_final.xlsx", index=True)

print('-------------Fim programa---------------')





# print('-------------início for---------------')    
# for i in range(0,46): 
#    print(nomeArquivos[i])
#    print(colecaoPlanilhas[i]['Aparelho gela?'])
#    print('-------------fim planilha---------------')
# print('-------------Fim programa---------------')    








# string = 'aparelho gela?'
# pattern = re.compile('.parelho .ela.')
# x = re.search(pattern,string)
# print(x)

#nomeArquivos = [""]


# wb = load_workbook('01.xlsx')
# wb.active
# abaColegio = wb['colegioaplicacao']

# print(abaColegio)


# planilha01 = pd.read_excel("01.xlsx")
# planilha02 = pd.read_excel("02.xlsx")
# planilha03 = pd.read_excel("03.xlsx")


# contagem_01 = (planilha01['Aparelho gela?'] == 'Não').sum()
# contagem_02 = (planilha02['Aparelho gela?'] == 'não').sum()
# contagem_03 = (planilha03['Aparelho gela?'] == 'não').sum()


# print(contagem_01)

# print(f'Planilha 01 - Aparelho Gela: {contagem_01} vezes')
# print(f'Planilha 02 - Aparelho Gela: {contagem_02} vezes')
# print(f'Planilha 03 - Aparelho Gela: {contagem_03} vezes')


# linha = planilha01['Aparelho gela?']
# print(linha[1])








#print(planilha01)