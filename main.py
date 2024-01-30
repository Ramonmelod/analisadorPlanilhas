import pandas as pd
#import PyPDF2
from openpyxl import load_workbook  # load_workbook permite abrir e carregar um arquivo Excel existente para manipulação.
import re
import os  # biblioteca usada para utilização do metodo de listagem do nome dos arquivos

caminhopasta = "C:\\Users\\ramon\\Desktop\\levantamento\\planilhas\\"  # Substitua pelo caminho da sua pasta

colecaoPlanilhas =[]

nomeArquivos = os.listdir(caminhopasta) # Lista todos os arquivos na pasta e armazena em uma array


for i in range(0, 46):
   #print(nomeArquivos[i])
   planilhaSetor = pd.read_excel(caminhopasta + nomeArquivos[i])
   colecaoPlanilhas.append(planilhaSetor)
    

print(colecaoPlanilhas[0]['Aparelho gela?'])



#nomeArquivos = [""]


# wb = load_workbook('01.xlsx')
# wb.active
# abaColegio = wb['colegioaplicacao']

# print(abaColegio)


# planilha01 = pd.read_excel("01.xlsx")
# planilha02 = pd.read_excel("02.xlsx")
# planilha03 = pd.read_excel("03.xlsx")


# contagem_01 = (planilha01['Aparelho gela?'] == 'Não').sum()
# contagem_02 = (planilha02['Aparelho gela?'] == 'não').sum()
# contagem_03 = (planilha03['Aparelho gela?'] == 'não').sum()


# print(contagem_01)

# print(f'Planilha 01 - Aparelho Gela: {contagem_01} vezes')
# print(f'Planilha 02 - Aparelho Gela: {contagem_02} vezes')
# print(f'Planilha 03 - Aparelho Gela: {contagem_03} vezes')


# linha = planilha01['Aparelho gela?']
# print(linha[1])








#print(planilha01)

